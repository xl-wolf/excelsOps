import openpyxl
import os
import json


# 创建文件夹
def mkdir(dir):
    if(os.path.exists(dir) == False):
        os.mkdir(dir)


finalResPath = './finalRes'
mkdir(finalResPath)


def readFromJson(file):
    with open(file, 'r', encoding='utf8') as fr:
        jsonData = json.load(fr)
    return jsonData


def setExcelHeader(json, sheet):
    idx = 0
    sheet_head = list(json[0].keys())
    while idx < len(sheet_head):
        sheet.cell(row=1, column=idx+1, value=sheet_head[idx])
        idx += 1


def setExcelBody(json, sheet):
    i = 0
    while i < len(json):
        row = json[i]
        cellValue = row['订单编号']
        sheet.cell(row=i + 2, column=1, value=cellValue)
        cellValue = row['收件人']
        sheet.cell(row=i + 2, column=2, value=cellValue)
        cellValue = row['收件人电话']
        sheet.cell(row=i + 2, column=3, value=cellValue)
        cellValue = row['收件人地址']
        sheet.cell(row=i + 2, column=4, value=cellValue)
        cellValue = row['发货状态']
        sheet.cell(row=i + 2, column=5, value=cellValue)
        cellValue = row['订单状态']
        sheet.cell(row=i + 2, column=6, value=cellValue)
        cellValue = row['渠道名称']
        sheet.cell(row=i + 2, column=7, value=cellValue)
        cellValue = row['商品名称']
        sheet.cell(row=i + 2, column=8, value=cellValue)
        cellValue = row['商品单价']
        sheet.cell(row=i + 2, column=9, value=cellValue)
        cellValue = row['商品数量']
        sheet.cell(row=i + 2, column=10, value=cellValue)
        cellValue = row['订单金额']
        sheet.cell(row=i + 2, column=11, value=cellValue)
        cellValue = row['应收金额']
        sheet.cell(row=i + 2, column=12, value=cellValue)
        cellValue = row['应付金额']
        sheet.cell(row=i + 2, column=13, value=cellValue)
        cellValue = row['售后渠道退款']
        sheet.cell(row=i + 2, column=14, value=cellValue)
        cellValue = row['售后仓库退款']
        sheet.cell(row=i + 2, column=15, value=cellValue)
        cellValue = row['利润']
        sheet.cell(row=i + 2, column=16, value=cellValue)
        i = i + 1


def writeToExcel(file):
    json = readFromJson(file)
    excel = openpyxl.Workbook()
    sheet = excel.create_sheet('明细', index=0)
    # 表头
    setExcelHeader(json, sheet)
    # 表主体
    setExcelBody(json, sheet)
    # 输出excel
    excel.save(finalResPath+'/最终要获得的数据.xlsx')


if __name__ == '__main__':
    writeToExcel('./mergedJson/最终要获得的数据.json')
