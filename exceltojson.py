import openpyxl
import json
import io
import os

# 全局变量
sourceDataDir = './sourceData'
tempJsonDir = './tempJson'


# excel表格转json文件
def excel_to_json(excel_file, active_sheet, json_f_name):
    jd = []
    heads = []
    book = openpyxl.load_workbook(excel_file)
    sheet = book[active_sheet]

    max_row = sheet.max_row
    max_column = sheet.max_column
    # 解析表头
    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row < 1:
            # 第一行跳过
            continue
        one_line = {}
        # 遍历一行中的每一个单元格
        for column in range(max_column):
            k = heads[column]
            v = sheet.cell(row + 1, column + 1).value
            one_line[k] = v
        jd.append(one_line)
    book.close()
    # 将json保存为文件
    save_json_file(jd, json_f_name)


# 创建文件夹
def mkdir(dir):
    if(os.path.exists(dir) == False):
        os.mkdir(dir)


# 读取文件数据
def read_json_config(path):
    dirs = []
    with open(path, 'r', encoding='utf-8') as fp:
        row_data = json.load(fp)
    # 读取每一条json数据
    for d in row_data:
        dirs = row_data[d]
    return dirs


# 将json保存为文件
def save_json_file(jd, json_f_name):
    mkdir(tempJsonDir)
    dirs = read_json_config('./dirconfig.json')
    for dir in dirs:
        mkdir(tempJsonDir+'/'+dir)
    f = io.open(json_f_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    f.write(txt)
    f.close()


# 入口方法
def entry_func():
    for path, dirs, files in os.walk(sourceDataDir):
        for file in files:
            # 需要处理的excel名称
            filename = path+'/'+file
            # 需要处理的excel表单名称
            sheetname = '明细'
            jsonPath = tempJsonDir + \
                path[len(sourceDataDir):]+'/'+file[0:-5]+'.json'
            excel_to_json(filename, sheetname, jsonPath)


if '__main__' == __name__:
    entry_func()
